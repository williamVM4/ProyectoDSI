from ast import If
from asyncio import current_task
from email.headerregistry import Group
from importlib.resources import contents
from inspect import _void
from multiprocessing import context
import re
'from tkinter import FALSE'
from urllib import response
from xml.dom.minidom import Identified
from django.forms import NullBooleanField
from django.http import request
from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.views.generic import CreateView, TemplateView,FormView, ListView, DetailView, UpdateView
from apps.monitoreo.models import estadoCuenta, condicionesPago
from apps.inventario.models import cuentaBancaria,proyectoTuristico, asignacionLote, detalleVenta
from apps.autenticacion.mixins import *
from .forms import *
from .models import *
from apps.inventario.models import *
from django.shortcuts import redirect, render
from django.contrib import messages
from crum import get_current_user
from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import *
from django.contrib.auth.models import User
from math import ceil, floor
from decimal import Decimal
from datetime import date, timedelta, time, datetime
from dateutil.relativedelta import relativedelta
from django.utils.dateparse import parse_datetime
import dateutil.parser

"Vista donde se listaran todos los pagos que se vayan registrando"
class caja(GroupRequiredMixin, FormView):
    group_required = [u'Configurador del sistema',u'Administrador del sistema',u'Operador del sistema']
    template_name = 'facturacion/caja.html'
    
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        "Se valida que exista el proyecto turistico"
        try:
            proyecto = proyectoTuristico.objects.get(pk=self.kwargs['idp'])
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el proyecto no existe')
            return HttpResponseRedirect(reverse_lazy('home'))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, *args, **kwargs):
        context=super().get_context_data(**kwargs)
        "Se envian los parametros por contexto y los objetos de tipo pago"
        id = self.kwargs.get('idp', None) 
        context['idp'] = id
        context['pagos'] = pago.objects.all().order_by('-fechaRegistro')
        return context

    def get_url_redirect(self, **kwargs):
        idp = self.kwargs.get('idp', None)
        return reverse_lazy('caja', kwargs={'idp': idp})

    form_class = resumenForm

    def form_valid(self, form, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"
        rPrima = form.cleaned_data.get("resumenPrima")
        rPagoM = form.cleaned_data.get("resumenPagoM")
        rPagoF = form.cleaned_data.get("resumenPagoF")
        fechaI = form.cleaned_data.get("fechaInicio")
        fechaF = form.cleaned_data.get("fechaFin")
        pagos = pago.objects.all()
        valido = False
        if fechaI <= fechaF:#Validar fechas
            rows = 2
            bordes = borders.Side(style = None, color = 'FF000000', border_style = 'thin')
            borde = Border(left = bordes, right = bordes, bottom = bordes, top = bordes)
            ws.merge_cells('B1:E1')
            ws['B1'] = 'Resumen de pagos'
            ws['B1'].alignment = Alignment(horizontal="center",vertical="center")
            ws['B2'] = 'Fecha'
            ws['C2'] = 'Lote'
            ws['D2'] = 'Monto'
            ws['E2'] = 'Concepto'

            if rPrima is True and rPagoM is True and rPagoF is True:
                pagos = pago.objects.filter()
            if rPrima is True and rPagoM is True and rPagoF is False: 
                pagos = pago.objects.filter(pagoFinanciamiento = None)
            if rPrima is True and rPagoM is False and rPagoF is True:
                pagos = pago.objects.filter(pagoMantenimiento = None)
            if rPrima is True and rPagoM is False and rPagoF is False:
                pagos = pago.objects.filter(pagoMantenimiento = None, pagoFinanciamiento = None)
            if rPrima is False and rPagoM is True and rPagoF is False:
                pagos = pago.objects.filter(prima = None, pagoFinanciamiento = None)
            if rPrima is False and rPagoM is True and rPagoF is True:
                pagos = pago.objects.filter(prima = None)
            if rPrima is False and rPagoM is False and rPagoF is True:
                pagos = pago.objects.filter(prima = None, pagoMantenimiento = None)
            if rPrima is False and rPagoM is False and rPagoF is False:
                messages.error(self.request, 'Seleccione el tipo de pago que desea imprimir')
                return HttpResponseRedirect(self.get_url_redirect())

            rows = 3
            totalPagos = 3
            for p in pagos:
                lote = ''
                concepto = ''
                if p.prima != None:
                    lote = p.prima.detalleVenta.lote.identificador
                    concepto = 'Prima'
                if p.pagoFinanciamiento != None:
                    lote = p.pagoFinanciamiento.estadoCuenta.detalleVenta.lote.identificador
                    concepto = 'Pago Financiamiento'
                if p.pagoMantenimiento != None:
                    lote = p.pagoMantenimiento.estadoCuenta.detalleVenta.lote.identificador
                    concepto = 'Pago Mantenimiento'
                if p.prima == None and p.pagoFinanciamiento == None and p.pagoMantenimiento == None:
                    messages.error(self.request, 'x')
                    return HttpResponseRedirect(self.get_url_redirect())
                if p.fechaPago >= fechaI and p.fechaPago <= fechaF:   
                    ws.cell(row=rows,column=2).value = p.fechaPago
                    ws.cell(row=rows,column=3).value = lote
                    ws.cell(row=rows,column=4).value = p.monto
                    ws.cell(row=rows,column=5).value = concepto
                    ws.cell(row=rows,column=4).number_format = '$ 0.00' 
                    rows = rows+1
                    totalPagos = totalPagos + 1
                    valido = True
            ws.cell(row=rows,column=4).number_format = '$ 0.00'               
            ws.cell(row=rows,column=4).value = ("=SUM(D"+str(3)+":D"+str(rows-1)+")") 
            ws.cell(row=rows,column=2).value = 'Total'  
            
            rows = 1
            for p in range(totalPagos):
                ws.cell(row=rows,column=2).border = borde
                ws.cell(row=rows,column=3).border = borde
                ws.cell(row=rows,column=4).border = borde
                ws.cell(row=rows,column=5).border = borde 
                rows = rows+1  
                      
        else: 
            messages.error(self.request, 'La fecha de fin no debe ser menor que la fecha inicio')
            return HttpResponseRedirect(self.get_url_redirect())
        if valido:
            nombre_archivo = "Resumen.xlsx"
            response = HttpResponse()
            contenido = "attachment; filename = {0}".format(nombre_archivo)
            response["Content-Disposition"]= contenido
            wb.save(response)
            return response
        else: 
            messages.error(self.request, 'No existen ingresos en ese intervalo de fechas')
            return HttpResponseRedirect(self.get_url_redirect())



"Vista de formulario para agregar prima"
class agregarPrima(GroupRequiredMixin,CreateView):
    group_required = [u'Configurador del sistema',u'Administrador del sistema',u'Operador del sistema']
    model = prima
    template_name = 'facturacion/Pago/Prima/agregarPrima.html'
    form_class = agregarPrimaForm #formulario con datos del modelo prima
    second_form_class = pagoForm #formulario con datos del modelo pago
    third_form_class = lotePagoForm #formulario con el detalle de venta

    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        "Se valida si el proyecto turistico existe"
        try:
            proyecto = proyectoTuristico.objects.get(pk=self.kwargs['idp'])
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el proyecto no existe')
            return HttpResponseRedirect(reverse_lazy('home'))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs): 
        context = super(agregarPrima, self).get_context_data(**kwargs)
        "Se envia el id del proyecto y los formularios auxiliares por contexto"
        idp = self.kwargs.get('idp', None)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(id = idp)
        if 'form3' not in context:
            context['form3'] = self.third_form_class(id = idp)
        context['idp'] = idp
        return context

    "url de exito del formulario"
    def get_url_redirect(self, **kwargs):
        idp = self.kwargs.get('idp', None)
        return reverse_lazy('caja', kwargs={'idp': idp})

    def form_valid(self, form, **kwargs):
        "Se obtienen los datos de los formularios"
        prima = form.save(commit=False)
        pago = self.second_form_class(self.request.POST).save(commit=False)
        lotef = self.third_form_class(self.request.POST).data['matricula']
        try:
            detalle = detalleVenta.objects.get(id = lotef)
            "validacion de que tenga propietarios"
            asig = asignacionLote.objects.filter(detalleVenta = detalle).exists()
            if asig is False:
                messages.error(self.request, 'Ocurrió un error, el lote '+detalle.lote.identificador+' no tiene propietarios registrados')
                return self.render_to_response(self.get_context_data(form=form, form2=self.second_form_class(self.request.POST), form3 = self.third_form_class(self.request.POST)))
            "validacion de que no tenga estado de cuenta"
            est = estadoCuenta.objects.filter(detalleVenta = detalle).exists()
            if est is True:
                messages.error(self.request, 'Ocurrió un error, el lote '+detalle.lote.identificador+' tiene un estado de cuenta generado. Ya no puede registrar más primas')
                return self.render_to_response(self.get_context_data(form=form, form2=self.second_form_class(self.request.POST), form3 = self.third_form_class(self.request.POST)))  
            pago.prima = prima
            pago.fechaRegistro=datetime.now()
            prima.detalleVenta = detalle
            "validacion de que el monto de prima no sea menor al precio de venta - descuento"
            if pago.monto > (detalle.precioVenta - detalle.descuento):
                messages.error(self.request, 'Ocurrió un error, el monto de la prima debe ser menor al precio de venta')
                return self.render_to_response(self.get_context_data(form=form, form2=self.second_form_class(self.request.POST), form3 = self.third_form_class(self.request.POST)))
            "guarda el usuario"
            user = get_current_user()
            if user is not None:
                prima.usuarioCreacion = user
            prima.save()
            pago.save()
            messages.success(self.request, 'La prima fue registrada con exito')
        except Exception:
            messages.error(self.request, 'Ocurrió un error al guardar la prima')
        return HttpResponseRedirect(self.get_url_redirect())

"Vista de formulario para consultar pago por mantenimiento"
class detallePago(GroupRequiredMixin,DetailView):
    group_required = [u'Configurador del sistema',u'Administrador del sistema']
    template_name = 'facturacion/Pago/detallePago.html'
    model = pago
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        """Validacion de que exista proyecto"""
        try:
            proyecto = proyectoTuristico.objects.get(pk=self.kwargs['idp'])
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el proyecto no existe')
            return HttpResponseRedirect(reverse_lazy('home'))
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        """Se recuperan los parametros necesarios pasados por url y se envia por contexto la lista
        de pagos por proyecto turistico"""
        id = self.kwargs.get('idp', None)
        idPago = self.kwargs.get('pk', None)
        context['idp'] = id
        context['pago'] = pago.objects.get(id = idPago)
        return context

"Vista de formulario para agregar pago por mantenimiento"
class agregarPagoMantenimiento(GroupRequiredMixin,CreateView):
    group_required = [u'Configurador del sistema',u'Administrador del sistema',u'Operador del sistema']
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        try:
            proyecto = proyectoTuristico.objects.get(pk=self.kwargs['idp'])
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el proyecto no existe')
            return HttpResponseRedirect(reverse_lazy('home'))
        return super().dispatch(request, *args, **kwargs)

    model = pagoMantenimiento
    form_class = agregarPagoMantenimientoForm
    second_form_class = pagoForm
    third_form_class = lotePagoForm
    template_name = 'facturacion/Pago/PagoMantenimiento/agregarPagoMantenimiento.html'

    def get_context_data(self, **kwargs):
        context = super(agregarPagoMantenimiento, self).get_context_data(**kwargs)
        idp = self.kwargs.get('idp', None)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(id = idp)
        if 'form3' not in context:
            context['form3'] = self.third_form_class(id = idp)
        context['idp'] = idp
        return context 
    
    def get_url_redirect(self, **kwargs):
        context=super().get_context_data(**kwargs)
        idp = self.kwargs.get('idp', None)
        return reverse_lazy('caja', kwargs={'idp': idp})

    def form_valid(self, form, **kwargs):
        context=super().get_context_data(**kwargs)
        
        lotef = self.third_form_class(self.request.POST).data['matricula']
        detalle = detalleVenta.objects.get(id = lotef)

        #Validacion de asignacion
        try:
             asig = asignacionLote.objects.filter(detalleVenta = detalle)
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el lote '+detalle.lote.identificador+' no tiene propietarios')
            return self.render_to_response(self.get_context_data(form=form)) 

        asig = asignacionLote.objects.filter(detalleVenta = detalle)

        #Validacion de estado de cuenta
        try:
            estaC = estadoCuenta.objects.get(detalleVenta = detalle)
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el lote '+detalle.lote.identificador+' no tiene un estado de cuenta generado. Establezca la venta para poder agregar un pago')
            return self.render_to_response(self.get_context_data(form=form))

        estadoC = estadoCuenta.objects.get(detalleVenta = detalle)
        condicion=condicionesPago.objects.get(detalleVenta=detalle)
        recibo=self.form_class(self.request.POST).data['numeroReciboMantenimiento']
        stringFechaPago=self.second_form_class(self.request.POST).data['fechaPago']
        stringFechaPago=stringFechaPago[0:4]+"-"+stringFechaPago[6:7]+"-"+stringFechaPago[8:10]
        fechaPago=dateutil.parser.parse(stringFechaPago).date()
        listadoPagos=pagoMantenimiento.objects.filter(estadoCuenta=estadoC).order_by('-fechaRegistro')
        primeraCuota="NO"
        saldoUltimaCuota=0
        saldoUltimoRecargo=0
        monto=0
        descuento=0
        montoOtros=0
        conceptoDescuento=""
        conceptoOtros=""
        observaciones=""
        
        #Verificar que existan cuotas de mantenimiento en el estado de cuenta
        try: 
            ultimoPagoListado=listadoPagos[0]
            ultimoPago=pago.objects.get(pagoMantenimiento = ultimoPagoListado)
        except Exception:
            primeraCuota="SI"
            pass

        #Validacion del monto
        try:
            monto=Decimal(self.second_form_class(self.request.POST).data['monto'])
            if monto==0:
                messages.error(self.request, 'Ocurrió un error, el monto no es un valor válido.')
                return self.render_to_response(self.get_context_data(form=form))  
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el monto no debe ser cero.')
            return self.render_to_response(self.get_context_data(form=form))

        #Obtenemos la fechas del corte, la cuota de pago, inicio y fin de mes
        fechaEscrituracion=condicion.fechaEscrituracion
        fechaCorte=condicion.fechaEscrituracion
        stringFechaCadaMes="FECHA DE PAGO: "+fechaCorte.strftime("%d")+" de cada mes.\n"
        fechaUltimoRecargo=fechaCorte
        stringMantenimiento=""
        stringRecargo=""
        stringDetalle=""
        tituloMtto="\nMANTENIMIENTO:\n"
        tituloRecargo="\nRECARGO:\n"
        valorPagado=monto
        fechaUltimoPago=fechaCorte
        calculoRercargo=0
        cantidadMeses=0

        #Validacion del monto en concepto de otros
        try:
            montoOtros=Decimal(self.form_class(self.request.POST).data['montoOtros'])
        except Exception:
            pass

        #Validacion del monto en concepto de descuento
        try:
            descuento=Decimal(self.form_class(self.request.POST).data['descuento'])
        except Exception:
            pass    

        #Si existe un monto en concepto de otros se lo restamos al valor pagado
        monto-=montoOtros
        valorRecargo=0
        abono=0

        #Calculo del ultimo pago ingresado
        if primeraCuota=="NO":
            fechaCorte=ultimoPago.pagoMantenimiento.fechaUltimoMtto
            saldoUltimaCuota=ultimoPago.pagoMantenimiento.abono
            fechaUltimoRecargo=ultimoPago.pagoMantenimiento.fechaUltimoRecargo
            saldoUltimoRecargo=ultimoPago.pagoMantenimiento.saldoRecargo
            fechaUltimoPago=ultimoPago.fechaPago
        
        #Validacion del concepto del descuento
        try:
            conceptoDescuento=self.form_class(self.request.POST).data['conceptoDescuento']
        except Exception:
            pass

        #Validacion del concepto de otros cargos
        try:
            conceptoOtros=self.form_class(self.request.POST).data['conceptoOtros']
        except Exception:
            pass

        #Validacion del concepto de las observaciones
        try:
            observaciones=self.second_form_class(self.request.POST).data['observaciones']
        except Exception:
            pass
            
        #Validacion de la fecha de pago
        if fechaPago<fechaUltimoPago:
            messages.error(self.request, 'Ocurrió un error, la fecha de pago debe ser mayor al último pago registrado en el estado de cuenta del lote. Fecha último pago '+fechaUltimoPago.strftime("%d/%m/%Y"))
            return self.render_to_response(self.get_context_data(form=form)) 
        
        #Validacion del concepto por descuento
        if descuento!=0 and conceptoDescuento=="":
            messages.error(self.request, 'Ocurrió un error, no se ha asignado un concepto al monto del descuento.')
            return self.render_to_response(self.get_context_data(form=form))

        #Validacion del concepto por otros pagos
        if montoOtros!=0 and conceptoOtros=="":
            messages.error(self.request, 'Ocurrió un error, no se ha asignado un concepto al monto por otros pagos.')
            return self.render_to_response(self.get_context_data(form=form))  

        #Validacion del monto a cancelar
        if montoOtros!=0 and descuento!=0:
            if valorPagado<(montoOtros+descuento):
                messages.error(self.request, 'Ocurrió un error, el monto no corresponde al valor a cancelar.')
                return self.render_to_response(self.get_context_data(form=form))
        elif montoOtros>valorPagado and descuento==0:
            messages.error(self.request, 'Ocurrió un error, el monto no corresponde al valor a cancelar en concepto de otros pagos.')
            return self.render_to_response(self.get_context_data(form=form))
        elif descuento>valorPagado and montoOtros==0:
            messages.error(self.request, 'Ocurrió un error, el monto no corresponde al valor a cancelar en concepto de descuento por recargo.')
            return self.render_to_response(self.get_context_data(form=form))

        #Validacion del monto de descuento, exista un recargo para cobrarlo
        fechaValidacion=fechaCorte + relativedelta(months=1)
        fechaValidacionActualizada=fechaActualizada(fechaValidacion,fechaEscrituracion)
        if  descuento !=0:
            if fechaPago>fechaValidacionActualizada:
                cantidadMeses=int(cantMeses(fechaPago,fechaUltimoRecargo))
                fechaRecargo=fechaUltimoRecargo+relativedelta(months=cantidadMeses)
                fechaRecargo=fechaActualizada(fechaRecargo, fechaEscrituracion)
                if fechaPago<=fechaRecargo:
                    cantidadMeses-=1
                montoMensajeRecargo=Decimal(cantidadMeses*condicion.multaMantenimiento)
                if descuento>montoMensajeRecargo:
                    messages.error(self.request, 'Ocurrió un error, el monto del descuento no debe ser mayor al cálculo del recargo. Recargo calculado: $ '+str(round(montoMensajeRecargo,2)))
                    return self.render_to_response(self.get_context_data(form=form)) 
            else:
                messages.error(self.request, 'Ocurrió un error, no existe un recargo al cual aplicar el descuento.')
                return self.render_to_response(self.get_context_data(form=form)) 
        
        #Calculo del mantenimiento 
        montoVeces=monto
        yaInicio=""
        cantVecesRC=0
        cantVecesRM=0
        descuentoInd=0
        if saldoUltimaCuota!=0:
            montoVeces-=(condicion.mantenimientoCuota-saldoUltimaCuota)
        if saldoUltimoRecargo!=0:
            montoVeces-=(condicion.multaMantenimiento-saldoUltimoRecargo)
        cantidadMesesPrint=int(cantMeses(fechaPago,fechaUltimoRecargo))
        fechaRecargoFin=fechaCorte + relativedelta(months=cantidadMesesPrint)
        fechaRecargoFin=fechaActualizada(fechaRecargoFin,fechaEscrituracion)
        if fechaPago<=fechaRecargoFin:
            cantidadMesesPrint-=1
            calculoRercargo=(cantidadMesesPrint*(condicion.mantenimientoCuota+condicion.multaMantenimiento))
            calculoRercargo=(montoVeces-calculoRercargo)
            calculoRercargoC=(cantidadMesesPrint*condicion.multaMantenimiento)
            cantVecesRC=floor(calculoRercargoC/condicion.multaMantenimiento)
            cantVecesRM=cantVecesRC+floor(calculoRercargo/(condicion.mantenimientoCuota+condicion.multaMantenimiento))
            descuentoInd=descuento/cantidadMesesPrint
        else:
            calculoRercargo=(cantidadMesesPrint*(condicion.mantenimientoCuota+condicion.multaMantenimiento))
            calculoRercargo=(montoVeces-calculoRercargo)
            calculoRercargoC=(cantidadMesesPrint*condicion.multaMantenimiento)
            cantVecesRC=floor(calculoRercargoC/condicion.multaMantenimiento)
            cantVecesRM=cantVecesRC+floor(calculoRercargo/(condicion.mantenimientoCuota+condicion.multaMantenimiento))
            descuentoInd=descuento/cantidadMesesPrint
        cantVeces=floor(montoVeces/condicion.mantenimientoCuota)
        cantVecesR=floor(montoVeces/(condicion.mantenimientoCuota+condicion.multaMantenimiento))
        i=-1
        while monto>0:
            i+=1
            fechaValidacion=fechaCorte + relativedelta(months=1)
            if fechaPago<=fechaActualizada(fechaValidacion,fechaEscrituracion):
                if saldoUltimoRecargo==0:
                    if saldoUltimaCuota==0:
                        if monto>=condicion.mantenimientoCuota:
                            fechaCorte=fechaCorte + relativedelta(months=1)
                            fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                            cantAntesVeces=valorPagado-montoVeces
                            if cantAntesVeces==0:
                                if cantVeces==1:
                                    stringMantenimiento+=printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota,2))+"\n"
                                else:
                                    if i == 0:
                                        stringMantenimiento+=printFecha(fechaCorte)+" a "
                                    elif i==(cantVeces-1) and yaInicio=="":
                                        stringMantenimiento+=printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota*cantVeces,2))+"\n"
                                    elif i==(cantVecesRM-1) and yaInicio!="":
                                        stringMantenimiento+=printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota*(cantVecesRM),2))+"\n"
                            else:
                                if cantVeces==1:
                                    stringMantenimiento+=printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota,2))+"\n"
                                else:
                                    if i==1:
                                        stringMantenimiento+=printFecha(fechaCorte)+" a "
                                    elif i==cantVeces and yaInicio=="":
                                        stringMantenimiento+=printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota*cantVeces,2))+"\n"
                                    elif i==cantVecesRM and yaInicio!="":
                                        stringMantenimiento+=printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota*cantVecesRM,2))+"\n"
                            registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Cuota "+printFecha(fechaCorte),condicion.mantenimientoCuota,0,0,0)
                            monto-=condicion.mantenimientoCuota
                        else:
                            fechaCorte=fechaCorte + relativedelta(months=1)
                            fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                            stringMantenimiento+="Ab. "+printFecha(fechaCorte)+"   $ "+str(round(monto,2))+"\nSaldo "+printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota-monto,2))+"\n"
                            registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Ab. Mantenimiento "+printFecha(fechaCorte),monto,0,0,0)
                            abono=monto
                            monto-=monto
                    else:
                        if monto>=(condicion.mantenimientoCuota-saldoUltimaCuota):
                            fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                            stringMantenimiento+="Compl. "+printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota-saldoUltimaCuota,2))+"\n"
                            registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Compl. Mantenimiento "+printFecha(fechaCorte),condicion.mantenimientoCuota-saldoUltimaCuota,0,0,0)
                            monto-=(condicion.mantenimientoCuota-saldoUltimaCuota)
                            abono=0
                            saldoUltimaCuota=0
                        else:
                            fechaCorte=fechaCorte + relativedelta(months=1)
                            fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                            stringMantenimiento+="Ab. "+printFecha(fechaCorte)+"   $ "+str(round(monto,2))+"\nSaldo "+printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota-(saldoUltimaCuota+monto),2))+"\n"
                            registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Ab. Mantenimiento "+printFecha(fechaCorte),monto,0,0,0)
                            abono=saldoUltimaCuota+monto
                            monto-=monto
                else:
                    if monto>=(condicion.multaMantenimiento-saldoUltimoRecargo):
                        fechaUltimoRecargo=fechaActualizada(fechaUltimoRecargo,fechaEscrituracion)
                        stringRecargo+="Compl. "+printFecha(fechaUltimoRecargo)+"   $ "+str(round(condicion.multaMantenimiento-saldoUltimoRecargo,2))+"\n"
                        registoCuotaM(recibo,estadoC,fechaPago,fechaUltimoRecargo,"Compl. Recargo "+printFecha(fechaUltimoRecargo),0,condicion.multaMantenimiento-saldoUltimoRecargo,0,0)
                        monto-=(condicion.multaMantenimiento-saldoUltimoRecargo)
                        saldoUltimoRecargo=0
                    else:
                        fechaUltimoRecargo=fechaUltimoRecargo + relativedelta(months=1)
                        fechaUltimoRecargo=fechaActualizada(fechaUltimoRecargo,fechaEscrituracion)
                        stringRecargo+="Ab. "+printFecha(fechaUltimoRecargo)+"   $ "+str(round(monto,2))+"\nSaldo "+printFecha(fechaUltimoRecargo)+"   $ "+str(round(condicion.multaMantenimiento-(saldoUltimoRecargo+monto),2))+"\n"
                        registoCuotaM(recibo,estadoC,fechaPago,fechaUltimoRecargo,"Ab. Recargo "+printFecha(fechaUltimoRecargo),0,condicion.multaMantenimiento-saldoUltimoRecargo,0,0)
                        saldoUltimoRecargo+=monto
                        monto-=monto
            else:
                if saldoUltimoRecargo==0:
                    if saldoUltimaCuota==0:
                        if monto>=(condicion.mantenimientoCuota+condicion.multaMantenimiento):
                            fechaCorte=fechaCorte + relativedelta(months=1)
                            fechaUltimoRecargo=fechaUltimoRecargo + relativedelta(months=1)
                            fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                            fechaUltimoRecargo=fechaActualizada(fechaUltimoRecargo,fechaEscrituracion)
                            cantAntesVeces=valorPagado-montoVeces
                            if cantAntesVeces==0:
                                if cantVecesR==1:
                                    stringMantenimiento+=printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota,2))+"\n"
                                    stringRecargo+=printFecha(fechaUltimoRecargo)+"   $ "+str(round(condicion.multaMantenimiento,2))+"\n"
                                else:
                                    if i == 0:
                                        stringMantenimiento+=printFecha(fechaCorte)+" a "
                                        stringRecargo+=printFecha(fechaUltimoRecargo)+" a "
                                        yaInicio="Si"
                                    elif i==(cantVecesR-1):
                                        stringMantenimiento+=printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota*cantVecesR,2))+"\n"
                                        stringRecargo+=printFecha(fechaUltimoRecargo)+"   $ "+str(round(condicion.multaMantenimiento*cantVecesR,2))+"\n"
                                    elif i==(cantVecesRC-1) and yaInicio!="":
                                        stringRecargo+=printFecha(fechaUltimoRecargo)+"   $ "+str(round((condicion.multaMantenimiento*cantVecesRC),2))+"\n"
                            else:
                                if cantVecesR==1:
                                    stringMantenimiento+=printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota,2))+"\n"
                                    stringRecargo+=printFecha(fechaUltimoRecargo)+"   $ "+str(round(condicion.multaMantenimiento,2))+"\n"
                                else:
                                    if i==1:
                                        stringMantenimiento+=printFecha(fechaCorte)+" a "
                                        stringRecargo+=printFecha(fechaUltimoRecargo)+" a "
                                        yaInicio="Si"
                                    elif i==cantVecesR:
                                        stringMantenimiento+=printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota*cantVecesR,2))+"\n"
                                        stringRecargo+=printFecha(fechaUltimoRecargo)+"   $ "+str(round(condicion.multaMantenimiento*cantVecesR,2))+"\n"
                                    elif i==cantVecesRC and yaInicio!="":
                                        stringRecargo+=printFecha(fechaUltimoRecargo)+"   $ "+str(round((condicion.multaMantenimiento*cantVecesRC),2))+"\n"                              
                            registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Cuota y Recargo "+printFecha(fechaCorte),condicion.mantenimientoCuota,condicion.multaMantenimiento,0,descuentoInd)
                            monto-=condicion.multaMantenimiento
                            valorRecargo+=condicion.multaMantenimiento
                            monto-=condicion.mantenimientoCuota
                            saldoUltimoRecargo=0
                            saldoUltimaCuota=0
                        else:
                            if fechaCorte == fechaUltimoRecargo:
                                if monto<condicion.multaMantenimiento:
                                    fechaUltimoRecargo=fechaUltimoRecargo + relativedelta(months=1)
                                    fechaUltimoRecargo=fechaActualizada(fechaUltimoRecargo,fechaEscrituracion)
                                    stringRecargo+="Ab. "+printFecha(fechaUltimoRecargo)+"   $ "+str(round(monto,2))+"\nSaldo "+printFecha(fechaUltimoRecargo)+"   $ "+str(round(condicion.multaMantenimiento-monto,2))+"\n"
                                    registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Ab. Recargo "+printFecha(fechaUltimoRecargo),0,monto,0,0)
                                    valorRecargo+=monto
                                    saldoUltimoRecargo=monto
                                    monto-=monto
                                elif monto>condicion.multaMantenimiento:
                                    fechaCorte=fechaCorte + relativedelta(months=1)
                                    fechaUltimoRecargo=fechaUltimoRecargo + relativedelta(months=1)
                                    fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                                    fechaUltimoRecargo=fechaActualizada(fechaUltimoRecargo,fechaEscrituracion)
                                    stringRecargo+=printFecha(fechaUltimoRecargo)+"   $ "+str(round(condicion.multaMantenimiento,2))+"\n"
                                    monto-=condicion.multaMantenimiento
                                    valorRecargo+=condicion.multaMantenimiento
                                    saldoUltimoRecargo=0
                                    stringMantenimiento+="Ab. "+printFecha(fechaCorte)+"   $ "+str(round(monto,2))+"\nSaldo "+printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota-monto,2))+"\n"
                                    registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Ab. Cuota y Recargo "+printFecha(fechaCorte),monto,condicion.multaMantenimiento,0,descuentoInd)
                                    abono+=monto
                                    monto-=monto
                                elif monto==condicion.multaMantenimiento:
                                    fechaUltimoRecargo=fechaUltimoRecargo + relativedelta(months=1)
                                    fechaUltimoRecargo=fechaActualizada(fechaUltimoRecargo,fechaEscrituracion)
                                    stringRecargo+=printFecha(fechaUltimoRecargo)+"   $ "+str(round(condicion.multaMantenimiento,2))+"\n"
                                    monto-=condicion.multaMantenimiento
                                    valorRecargo+=condicion.multaMantenimiento
                                    saldoUltimoRecargo=0
                                    registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Recargo "+printFecha(fechaUltimoRecargo),0,condicion.multaMantenimiento,0,0)
                            elif fechaCorte<fechaUltimoRecargo:
                                if saldoUltimaCuota==0:
                                    if monto>=condicion.mantenimientoCuota:
                                        fechaCorte=fechaCorte + relativedelta(months=1)
                                        fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                                        stringMantenimiento+=printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota,2))+"\n"
                                        registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Cuota "+printFecha(fechaCorte),condicion.mantenimientoCuota,0,0,0)
                                        monto-=condicion.mantenimientoCuota
                                        abono=0
                                        saldoUltimaCuota=0
                                    else:
                                        fechaCorte=fechaCorte + relativedelta(months=1)
                                        fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                                        stringMantenimiento+="Ab. "+printFecha(fechaCorte)+"   $ "+str(round(monto,2))+"\nSaldo "+printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota-(saldoUltimaCuota+monto),2))+"\n"
                                        registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Ab. Mantenimiento "+printFecha(fechaCorte),monto,0,0,0)
                                        abono=monto
                                        saldoUltimaCuota=monto
                                        monto-=monto
                                else:
                                    if monto>=(condicion.mantenimientoCuota-saldoUltimaCuota):
                                        fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                                        stringMantenimiento+="Compl. "+printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota-saldoUltimaCuota,2))+"\n"
                                        registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Compl. Mantenimiento "+printFecha(fechaCorte),condicion.mantenimientoCuota-saldoUltimaCuota,0,0,0)
                                        monto-=(condicion.mantenimientoCuota-saldoUltimaCuota)
                                        abono=0
                                        saldoUltimaCuota=0
                                    else:
                                        fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                                        stringMantenimiento+="Ab. "+printFecha(fechaCorte)+"   $ "+str(round(monto,2))+"\nSaldo "+printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota-(saldoUltimaCuota+monto),2))+"\n"
                                        registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Ab. Mantenimiento "+printFecha(fechaCorte),monto,0,0,0)
                                        abono=saldoUltimaCuota+monto
                                        saldoUltimaCuota+=monto
                                        monto-=monto
                    else:
                        if monto>=(condicion.mantenimientoCuota-saldoUltimaCuota):
                            fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                            stringMantenimiento+="Compl. "+printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota-saldoUltimaCuota,2))+"\n"
                            registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Compl. Mantenimiento "+printFecha(fechaCorte),condicion.mantenimientoCuota-saldoUltimaCuota,0,0,0)
                            monto-=(condicion.mantenimientoCuota-saldoUltimaCuota)
                            saldoUltimaCuota=0
                        else:
                            fechaCorte=fechaActualizada(fechaCorte,fechaEscrituracion)
                            stringMantenimiento+="Ab. "+printFecha(fechaCorte)+"   $ "+str(round(monto,2))+"\nSaldo "+printFecha(fechaCorte)+"   $ "+str(round(condicion.mantenimientoCuota-(saldoUltimaCuota+monto),2))+"\n"
                            registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Ab. Mantenimiento "+printFecha(fechaCorte),monto,0,0,0)
                            abono=saldoUltimaCuota+monto
                            monto-=monto
                else:
                    if monto>=(condicion.multaMantenimiento-saldoUltimoRecargo):
                        fechaUltimoRecargo=fechaActualizada(fechaUltimoRecargo,fechaEscrituracion)
                        stringRecargo+="Compl. "+printFecha(fechaUltimoRecargo)+"   $ "+str(round(condicion.multaMantenimiento-saldoUltimoRecargo,2))+"\n"
                        registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Compl. Recargo "+printFecha(fechaUltimoRecargo),0,condicion.multaMantenimiento-saldoUltimoRecargo,0,0)
                        valorRecargo+=(condicion.multaMantenimiento-saldoUltimoRecargo)
                        saldoUltimoRecargo=0
                        monto-=(condicion.multaMantenimiento-saldoUltimoRecargo)
                    else:
                        fechaUltimoRecargo=fechaActualizada(fechaUltimoRecargo,fechaEscrituracion)
                        stringRecargo+="Ab. "+printFecha(fechaUltimoRecargo)+"   $ "+str(round(monto,2))+"\nSaldo "+printFecha(fechaUltimoRecargo)+"   $ "+str(round(condicion.multaMantenimiento-(saldoUltimoRecargo+monto),2))+"\n"
                        registoCuotaM(recibo,estadoC,fechaPago,fechaCorte,"Ab. Recargo "+printFecha(fechaUltimoRecargo),0,monto,0,0)
                        valorRecargo+=monto
                        saldoUltimoRecargo+=monto
                        monto-=monto

        #Actualizamos el abono de la cuota de acuerdo a lo cancelado
        if valorPagado==montoOtros:
            fechaCorte=ultimoPago.pagoMantenimiento.fechaUltimoMtto
            saldoUltimaCuota=ultimoPago.pagoMantenimiento.abono
            fechaUltimoRecargo=ultimoPago.pagoMantenimiento.fechaUltimoRecargo
            saldoUltimoRecargo=ultimoPago.pagoMantenimiento.saldoRecargo
            abono=saldoUltimaCuota
            valorPagado=0
            valorRecargo=0

        #Actualizamos la fecha del recargo de acuerdo a lo cancelado
        if saldoUltimoRecargo==0:
            if fechaUltimoRecargo<=fechaCorte:
                if fechaPago<=fechaCorte:
                    fechaUltimoRecargo=fechaCorte
        #Imprimir lo calculado de mantenimiento
        if stringMantenimiento!="":
            stringDetalle+=stringFechaCadaMes+tituloMtto+stringMantenimiento

        #Imprimir lo calculado de recargo
        if stringRecargo!="":
            if stringMantenimiento!="":
                stringDetalle+=tituloRecargo+stringRecargo
            else:
                stringDetalle+=stringFechaCadaMes+tituloRecargo+stringRecargo
        #Imprimir el monto de otros si existe
        if montoOtros!=0:
            if montoOtros==valorPagado:
                stringDetalle+="OTROS:\n"+conceptoOtros+"  $ "+str(round(montoOtros,2))+"\n"
            else:
                stringDetalle+="\nOTROS:\n"+conceptoOtros+"   $ "+str(round(montoOtros,2))+"\n"
            registoCuotaM(recibo, estaC, fechaPago, fechaPago, "Otros: "+conceptoOtros, 0,0,montoOtros,0)
        
        #Imprimir el descuento si existe
        if descuento!=0:
            stringDetalle+="\nDESCUENTO DE RECARGO:    $ "+str(round(descuento,2))+"\n"+conceptoDescuento+"\n"            
        
        #Imprimir las observaciones si existen
        if observaciones!="":
            stringDetalle+="\nOBSERVACIONES:\n"+observaciones

        #Guardamos el pago de mantenimiento en la base de datos
        pagoM = form.save(commit=False)
        pagoF = self.second_form_class(self.request.POST).save(commit=False)
        try:
            pagoM.estadoCuenta = estadoC
            pagoM.fechaUltimoMtto=fechaActualizada(fechaCorte,fechaEscrituracion)
            pagoM.abono=abono
            pagoM.fechaUltimoRecargo=fechaActualizada(fechaUltimoRecargo,fechaEscrituracion)
            pagoM.saldoRecargo=saldoUltimoRecargo
            pagoM.mantenimiento=valorPagado-valorRecargo-montoOtros
            pagoM.recargoMtto=valorRecargo-descuento
            pagoM.conceptoDescuento=conceptoDescuento
            pagoM.descuento=descuento
            pagoM.conceptoOtros=conceptoOtros
            pagoM.montoOtros=montoOtros
            user = get_current_user()
            if user is not None:
                pagoM.usuarioCreacion = user
            pagoM.save()
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el lote '+detalle.lote.identificador+' no se le ha establecido una venta. Establezca las condiciones de pago del lote para poder agregar un pago')
            return self.render_to_response(self.get_context_data(form=form))   
        pagoF.pagoMantenimiento = pagoM
        pagoF.monto = valorPagado
        pagoF.observaciones = stringDetalle
        if pagoF.tipoPago == 1:
            pagoF.referencia = ''
            pagoF.cuentaBancaria = None
        pagoF.save()
        return HttpResponseRedirect(self.get_url_redirect())

#Función que registra la cuota de mantenimiento
def registoCuotaM(recibo, estadoC, fechaP, fechaC, conceptoC, mantenimientoC, recargoC, montoO, descuentoC):
    pagoCuotaM=pagoCuotaMantenimiento(numeroReciboMantenimiento=recibo,estadoCuenta=estadoC, fechaRegistro=datetime.now(),fechaPago=fechaP, fechaCorte=fechaC, concepto=conceptoC, mantenimiento=mantenimientoC, recargo=recargoC, otros=montoO,descuento=descuentoC)
    pagoCuotaM.save()
    return _void

#Función que calcula la cantidad de meses entre dos meses
def cantMeses(fechaMayor, fechaMenor):
    return (fechaMayor.year - fechaMenor.year) * 12 + fechaMayor.month - fechaMenor.month

#Función que valida el día de la fecha
def fechaActualizada(fecha, fechaEscrituracion):
    diaFecha=int(fecha.strftime('%d'))
    mesFecha=int(fecha.strftime('%m'))
    diaFechaEscrituracion=int(fechaEscrituracion.strftime('%d'))
    if diaFecha == diaFechaEscrituracion:
        return fecha
    elif diaFechaEscrituracion==29 and mesFecha==2:
        fecha=fecha+relativedelta(day=31)
        return fecha
    elif diaFechaEscrituracion==30 and mesFecha==2:
        fecha=fecha+relativedelta(day=31)
        return fecha
    elif diaFechaEscrituracion==31:
        fecha=fecha+relativedelta(day=31)
        return fecha
    elif diaFecha<diaFechaEscrituracion:
        diferencia=diaFechaEscrituracion-diaFecha
        fecha=fecha+relativedelta(days=diferencia)
        return fecha
    else:
        diferencia=diaFecha-diaFechaEscrituracion
        fecha=fecha-relativedelta(days=diferencia)
        return fecha

#Función que convierte fecha en formato español
def printFecha(fecha):
    mes=int(fecha.strftime('%m'))
    stringFecha=""
    if mes==1:
        stringFecha="Enero/"+fecha.strftime('%y') 
    elif mes==2:
        stringFecha="Febrero/"+fecha.strftime('%y')
    elif mes==3:
        stringFecha= "Marzo/"+fecha.strftime('%y')
    elif mes==4:
        stringFecha="Abril/"+fecha.strftime('%y')
    elif mes==5:
        stringFecha="Mayo/"+fecha.strftime('%y')
    elif mes==6:
        stringFecha="Junio/"+fecha.strftime('%y')
    elif mes==7:
        stringFecha="Julio/"+fecha.strftime('%y')
    elif mes==8:
        stringFecha="Agosto/"+fecha.strftime('%y')
    elif mes==9:
        stringFecha="Septiembre/"+fecha.strftime('%y')
    elif mes==10:
        stringFecha="Octubre/"+fecha.strftime('%y')
    elif mes==11:
        stringFecha="Noviembre/"+fecha.strftime('%y')
    elif mes==12:
        stringFecha="Diciembre/"+fecha.strftime('%y')
    return stringFecha
        
#--------------------Views de cuentas Bancarias----------------------------
"Vista de la lista de cuentas bancarias"
class gestionarCuentasBancarias(GroupRequiredMixin,ListView):
    group_required = [u'Configurador del sistema',u'Administrador del sistema']
    template_name = 'facturacion/CuentasBancarias/gestionarCuentasBancarias.html'
    model = cuentaBancaria
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        "Validacion de que exista el proyecto"
        try:
            proyecto = proyectoTuristico.objects.get(pk=self.kwargs['idp'])
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el proyecto no existe')
            return HttpResponseRedirect(reverse_lazy('home'))
        return super().dispatch(request, *args, **kwargs)
    
    "Se envia el parametro del id del proyecto por url y la lista de cuentas bancarias"
    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        id = self.kwargs.get('idp', None) 
        context['idp'] = id
        context['cuentaBancaria'] = cuentaBancaria.objects.filter(proyectoTuristico__id=id)
        return context

"Vista del formulario para agregar cuentas bancarias"
class agregarCuentaBancaria(GroupRequiredMixin,CreateView):
    group_required = [u'Configurador del sistema',u'Administrador del sistema']
    template_name = 'facturacion/CuentasBancarias/agregarCuentaBancaria.html'
    form_class = agregarCuentaBancariaForm
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        "Validacion de que exista el proyecto turistico"
        try:
            proyecto = proyectoTuristico.objects.get(pk=self.kwargs['idp'])
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el proyecto no existe')
            return HttpResponseRedirect(reverse_lazy('home'))
        return super().dispatch(request, *args, **kwargs)
    
    "Metodo para obtener la url de exito"
    def get_url_redirect(self, **kwargs):
        idp = self.kwargs.get('idp', None) 
        return reverse_lazy('cuentas', kwargs={'idp': idp})

    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        idp = self.kwargs.get('idp', None)
        context['idp'] = idp         
        return context

    def form_valid(self, form, **kwargs):
        idp = self.kwargs.get('idp', None) 
        cuentaBancaria = form.save(commit=False)
        try:
            cuentaBancaria.proyectoTuristico = proyectoTuristico.objects.get(id=idp)
            cuentaBancaria.identificador = str(cuentaBancaria.numeroCuentaBancaria)
            cuentaBancaria.save()
            messages.success(self.request, 'La cuenta bancaria se ha guardado con éxito')
        except Exception:
            messages.error(self.request, 'Ocurrió un error al guardar la cuenta bancaria, la cuenta bancaria no es válido')
        return HttpResponseRedirect(self.get_url_redirect())
    
#Eliminar prima 
class EliminarPrima(GroupRequiredMixin,TemplateView):
    group_required = [u'Configurador del sistema',u'Administrador del sistema',u'Operador del sistema']
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        try:
            proyecto = proyectoTuristico.objects.get(pk=self.kwargs['idp'])
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el proyecto no existe')
            return HttpResponseRedirect(reverse_lazy('home'))
        return super().dispatch(request, *args, **kwargs)
    
    def get_url_redirect(self, **kwargs):
        idp = self.kwargs.get('idp', None) 
        id = self.kwargs.get('idv', None) 
        try:
            detalleVenta.objects.get(pk = id)
            return reverse_lazy('detalleLote', kwargs={'idp': idp, 'pk': id})
        except Exception:
            return reverse_lazy('gestionarLotes', kwargs={'idp': idp})

    def get(self,request,*args,**kwargs):
        id = self.kwargs.get('id', None)
        try:
            primas = prima.objects.get(numeroReciboPrima = id)
            pagos = pago.objects.get(prima_id = id)
            try:
                condiciones = condicionesPago.objects.get(detalleVenta_id = primas.detalleVenta_id)
                messages.error(self.request, 'Error al eliminar la cuotra de prima, este lote ya cuenta con condiciones de pago establecidas.')
            except Exception:
                primas.delete() 
                pagos.delete()
                messages.success(self.request, 'La prima fue eliminada con exito.')      
        except Exception: 
            messages.error(self.request, 'Ocurrió un error al eliminar la prima, la prima no existe.')
        return HttpResponseRedirect(self.get_url_redirect())
        
#Modificar Prima 
class ModificarPrima(GroupRequiredMixin, UpdateView):
    group_required = [u'Configurador del sistema',u'Administrador del sistema',u'Operador del sistema']
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        try:
            proyecto = proyectoTuristico.objects.get(pk=self.kwargs['idp'])  
            try:
                detallev= detalleVenta.objects.get(id = self.kwargs['idv'])
                try:
                    condicion = condicionesPago.objects.get(detalleVenta_id = detallev.id)
                    messages.error(self.request, 'Ocurrió un error. Ya existe condiciones de pago, no es posible modificar la cuota de prima')
                    return HttpResponseRedirect(reverse_lazy('detalleLote', kwargs={'idp': self.kwargs['idp'], 'pk': self.kwargs['idv']}))
                except Exception:
                    try:
                        primas = prima.objects.get(numeroReciboPrima=self.kwargs['pk'])
                        if primas.detalleVenta != detallev:
                            messages.error(self.request, 'Ocurrió un error, la cuota de prima no pertenece a el detalle de venta')
                            return HttpResponseRedirect(reverse_lazy('detalleLote', kwargs={'idp': self.kwargs['idp'], 'pk': self.kwargs['idv']}))
                    except Exception:
                        messages.error(self.request, 'Ocurrió un error, el recibo de prima no existe')
                        return HttpResponseRedirect(reverse_lazy('detalleLote', kwargs={'idp': self.kwargs['idp'], 'pk': self.kwargs['idv']}))
            except Exception:
                messages.error(self.request, 'Ocurrió un error, el detalle de venta no existe')
                return HttpResponseRedirect(reverse_lazy('home'))
        except Exception:
            messages.error(self.request, 'Ocurrió un error, el proyecto no existe')
            return HttpResponseRedirect(reverse_lazy('home'))
        return super().dispatch(request, *args, **kwargs)

    model = prima
    second_model = pago
    third_model = lote
    template_name = 'facturacion/Pago/Prima/agregarPrima.html'
    form_class = agregarPrimaForm
    second_form_class = pagoForm
    third_form_class = lotePagoForm
    
    def get_context_data(self, **kwargs):
        context = super(ModificarPrima, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk', None)
        primas = self.model.objects.get(numeroReciboPrima = pk)
        pagos = self.second_model.objects.get(prima_id = primas.numeroReciboPrima)        
        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance = pagos)
        context['pk'] = pk   
        return context

    def get_form(self, form_class = None, **kwargs):
        form = super().get_form(form_class)
        form.fields['numeroReciboPrima'].disabled = True        
        return form

    def post(self, request, form_class = None, *args, **kwargs):
        self.object = self.get_object
        id_prima = kwargs['pk']       
        primas = self.model.objects.get(numeroReciboPrima = id_prima)
        pagos = self.second_model.objects.get(prima_id = primas.numeroReciboPrima)
        form = self.form_class(request.POST, instance = primas)
        form.fields['numeroReciboPrima'].disabled = True 
        
        form2 = self.second_form_class(request.POST, instance = pagos)
        idp = self.kwargs.get('idp', None) 
        id = self.kwargs.get('idv', None) 
        if form.is_valid() and form2.is_valid():
            primasF = form.save(commit = False)
            pagosF = form2.save(commit = False)          
            if pagosF.tipoPago == 1:
                pagosF.referencia = ''
                pagosF.cuentaBancaria = None

            primasF.save()
            pagosF.save()
            messages.success(self.request, 'Prima actualizada exitosamente')
        else: 
            messages.error(self.request, 'Ocurrió un error, el recibo de prima no se actualizo')
        return HttpResponseRedirect(reverse_lazy('detalleLote', kwargs={'idp': idp, 'pk': id}))  
       
    
class Recibo(TemplateView):

    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        idp = self.kwargs.get('idp', None) 
        id = self.kwargs.get('pk', None)
        context['idp'] = idp
        context['id'] = id
        return context

    
    def get(self,request,*args,**kwargs):
        context=super().get_context_data(**kwargs)
         # recojo el parametro 
        idp = self.kwargs.get('idp', None) 
        id = self.kwargs.get('pk', None)
        pagoRecibo = pago.objects.get(pk = id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Recibo"

        if pagoRecibo.prima_id:
            #tamanio columnas
            ws.column_dimensions['A'].width = 9.57
            ws.column_dimensions['B'].width = 13
            ws.column_dimensions['C'].width = 2.71
            ws.column_dimensions['D'].width = 11.14
            ws.column_dimensions['E'].width = 15.43
            ws.column_dimensions['F'].width = 11
            ws.column_dimensions['G'].width = 20.29
            ws.column_dimensions['H'].width = 9.71
            ws.column_dimensions['I'].width = 0
            ws.row_dimensions[1].height = 24.75
            ws.row_dimensions[2].height = 20.25
            ws.row_dimensions[3].height = 14
            ws.row_dimensions[4].height = 18.75
            ws.row_dimensions[5].height = 35.5
            ws.row_dimensions[6].height = 17.25
            ws.row_dimensions[7].height = 15
            ws.row_dimensions[8].height = 17.25
            ws.row_dimensions[9].height = 13.5
            ws.row_dimensions[10].height = 16.5
            ws.row_dimensions[11].height = 15.75
            ws.row_dimensions[12].height = 12
            ws.row_dimensions[13].height = 10.5
            ws.row_dimensions[14].height = 10.5
            ws.row_dimensions[15].height = 14.25
            ws.row_dimensions[16].height = 15.75
            ws.row_dimensions[17].height = 15.75
            ws.row_dimensions[18].height = 18
            ws.row_dimensions[19].height = 24
            ws['F5'] = 0
            ws['F7'] = 0
            ws['F8'] = 0
            ws['F9'] = 0
            ws.merge_cells('B2:F2')
            ws.merge_cells('B15:F15')
            ws['F5'].number_format = '0.00'
            ws['F7'].number_format = '0.00'
            ws['F8'].number_format = '0.00'
            ws['F9'].number_format = '0.00'
            ws['F10'].number_format = '0.00'  
            ws['F11'].number_format = '0.00'
            ws['B1'].number_format = '0.00'
            primaRecibo = prima.objects.get(numeroReciboPrima = pagoRecibo.prima_id )
            detalle = detalleVenta.objects.get(id = primaRecibo.detalleVenta_id)
            asigna = asignacionLote.objects.filter(detalleVenta_id = detalle.id)
            nombre = ""
            cantidad = 1
            for asignacion in asigna:
                if cantidad == 1:
                    nombre = asignacion.propietario.nombrePropietario
                else:
                    if cantidad <= 2:
                        nombre = nombre + " y Otros" 
                cantidad = cantidad +1

            lotes = lote.objects.get(matriculaLote = detalle.lote_id)
            usuario = User.objects.get(id = primaRecibo.usuarioCreacion_id)
            ws['G1'] = "Nº "+primaRecibo.numeroReciboPrima  
            ws['B2'] = nombre
            ws['B2'].alignment = Alignment(horizontal="center",vertical="center")
            ws['B4'].alignment = Alignment(horizontal="center",vertical="center")
            ws['C4'].alignment = Alignment(horizontal="center",vertical="center")
            ws['B4'] = lotes.numeroLote
            ws['C4'] = lotes.poligono
            ws['F10'] = pagoRecibo.monto
            ws['F11'] = "=SUM(F5:F10)"
            ws['B19'] = usuario.first_name
            ws['B1'] = '=F11'
            fecha = str(pagoRecibo.fechaPago).split(sep='-')
            f = reversed(fecha)
            
            for x in f:
                if ws.cell(row=15,column=2).value is None:
                    ws.cell(row=15,column=2).value = x 
                else:
                    ws.cell(row=15,column=2).value = str(ws.cell(row=15,column=2).value) +"                    "+x  
            ws['B15'].alignment = Alignment(horizontal="right",vertical="center")
            if pagoRecibo.tipoPago == 1:
                    ws['E18'] = "Pago realizado en efectivo"
            else:
                banco = cuentaBancaria.objects.get(numeroCuentaBancaria = pagoRecibo.cuentaBancaria_id)
                ws['E18'] = "Pago realizado en "+banco.banco+" Cta. No. "+banco.numeroCuentaBancaria
                ws['E19'] = "a/n de Decatur, S.A. Ref. " + pagoRecibo.referencia
        else:
            if pagoRecibo.pagoMantenimiento_id:
                ws.column_dimensions['A'].width = 9.43
                ws.column_dimensions['B'].width = 9.71
                ws.column_dimensions['C'].width = 1.29
                ws.column_dimensions['D'].width = 10.71
                ws.column_dimensions['E'].width = 14.14
                ws.column_dimensions['F'].width = 12.14
                ws.column_dimensions['G'].width = 10.71
                ws.column_dimensions['H'].width = 15.43
                ws.column_dimensions['I'].width = 10.71
                ws.row_dimensions[1].height = 11.75
                ws.row_dimensions[2].height = 12.5
                ws.row_dimensions[3].height = 27.75
                ws.row_dimensions[4].height = 18
                ws.row_dimensions[5].height = 15.75
                ws.row_dimensions[6].height = 18.75
                ws.row_dimensions[7].height = 15.75
                ws.row_dimensions[8].height = 18.75
                ws.row_dimensions[9].height = 17.25
                ws.row_dimensions[10].height = 15.75
                ws.row_dimensions[11].height = 21
                ws.row_dimensions[12].height = 15.75
                ws.row_dimensions[13].height = 22.5
                ws.row_dimensions[14].height = 15.75
                ws.row_dimensions[15].height = 15
                ws.row_dimensions[16].height = 18
                ws.row_dimensions[17].height = 15.75
                ws.row_dimensions[18].height = 15.75
                ws.row_dimensions[19].height = 18
                ws.row_dimensions[20].height = 21.75
                ws.row_dimensions[21].height = 12.75
                ws['F7'].number_format = ' 0.00'
                ws['F8'].number_format = ' 0.00'
                ws['F9'].number_format = ' 0.00'
                ws['F10'].number_format = ' 0.00'
                ws['F11'].number_format = ' 0.00'  
                ws['F12'].number_format = ' 0.00'
                ws['F13'].number_format = ' 0.00'
                ws['B3'].number_format = ' 0.00'
                ws['F7'] = 0
                ws['F8'] = 0
                ws['F9'] = 0
                ws['F10'] = 0
                ws['F11'] = 0
                ws['F12'] = 0
                ws['F13'] = 0
                ws['B3'] = 0
                pagoMRecibo = pagoMantenimiento.objects.get(numeroReciboMantenimiento = pagoRecibo.pagoMantenimiento_id)
                usuario = User.objects.get(id = pagoMRecibo.usuarioCreacion_id)
                estadoC = estadoCuenta.objects.get(id = pagoMRecibo.estadoCuenta_id)
                detalle = detalleVenta.objects.get(id = estadoC.detalleVenta_id)
                asigna = asignacionLote.objects.filter(detalleVenta_id = detalle.id)
                nombre = ""
                cantidad = 1
                for asignacion in asigna:
                    if cantidad == 1:
                        nombre = asignacion.propietario.nombrePropietario
                    else:
                        if cantidad <= 2:
                            nombre = nombre + " y Otros" 
                    cantidad = cantidad +1
                lotes = lote.objects.get(matriculaLote = detalle.lote_id)
                ws['H3'] = "Nº "+pagoMRecibo.numeroReciboMantenimiento
                ws.merge_cells('B4:F4')
                ws.merge_cells('B16:F16')
                ws.merge_cells('B12:E12')
                ws['F7'] = pagoRecibo.monto
                ws['F13'] = '=SUM(F7:F12)'
                ws['B3'] = '=F13'
                ws['B20'] = usuario.first_name
                ws['E18'].font = Font(size=10)
                ws['E19'].font = Font(size=10)
                fecha = str(pagoRecibo.fechaPago).split(sep='-')
                f = reversed(fecha)
                
                for x in f:
                    if ws.cell(row=16,column=2).value is None:
                        ws.cell(row=16,column=2).value = x 
                    else:
                        ws.cell(row=16,column=2).value = str(ws.cell(row=16,column=2).value) +"                    "+x  
                ws['B16'].alignment = Alignment(horizontal="right",vertical="center")


                
                ws['F12'] = pagoMRecibo.montoOtros
                ws['B12'] = "  "+pagoMRecibo.conceptoOtros
                ws['B4'].alignment = Alignment(horizontal="center",vertical="center")
                ws['B4'] = nombre
                ws['D5'].alignment = Alignment(horizontal="center",vertical="center")
                ws['E5'].alignment = Alignment(horizontal="center",vertical="center")
                ws['D5'] = lotes.numeroLote
                ws['E5'] = lotes.poligono

                if pagoRecibo.tipoPago == 1:
                    ws['E18'] = "Pago realizado en efectivo"
                else:
                    banco = cuentaBancaria.objects.get(numeroCuentaBancaria = pagoRecibo.cuentaBancaria_id)
                    ws['E18'] = "Pago realizado en "+banco.banco+" Cta. No. "+banco.numeroCuentaBancaria
                    ws['E19'] = "a/n de Decatur, S.A. Ref. " + pagoRecibo.referencia
                

        nombre_archivo = "Recibo.xlsx"
        response = HttpResponse()
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"]= contenido
        wb.save(response)
        return response
    
    

    