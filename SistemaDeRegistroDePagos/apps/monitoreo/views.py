from django.http import HttpResponseRedirect
from django.urls import reverse_lazy
from django.views.generic import ListView,TemplateView
from django.contrib.auth import login
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from apps.inventario.models import asignacionLote, detalleVenta, proyectoTuristico
from apps.monitoreo.models import estadoCuenta, cuotaEstadoCuenta
from apps.facturacion.models import pago, pagoMantenimiento
from apps.autenticacion.mixins import *
from django.contrib import messages
from .forms import *
from openpyxl import Workbook
from openpyxl.styles import *
from django.http.response import HttpResponse
from django.db.models import Sum

class estadoCuentaView(GroupRequiredMixin,ListView):
    group_required = [u'Configurador del sistema',u'Administrador del sistema']
    @method_decorator(login_required)
    def dispatch(self, request, *args, **kwargs):
        try:
            proyecto = proyectoTuristico.objects.get(pk=self.kwargs['idp'])
        except Exception:
            messages.error(self.request, 'Ocurrió un error, asegurese de que el proyecto existe')
            return HttpResponseRedirect(reverse_lazy('home'))
        try:
            lot = detalleVenta.objects.get(pk=self.kwargs['pk'])
        except Exception:
            messages.error(self.request, 'Ocurrió un error, asegurese de que el detalle de la venta existe')
            return HttpResponseRedirect(reverse_lazy('gestionarLotes', kwargs={'idp': self.kwargs['idp']}))
        try:
            estado = estadoCuenta.objects.get(detalleVenta=self.kwargs['pk'])
        except Exception:
            messages.error(self.request, 'Ocurrió un error el lote no tiene estado de cuenta generado. Verifique que ingreso las condiciones de pago')
            return HttpResponseRedirect(reverse_lazy('detalleLote', kwargs={'idp': self.kwargs['idp'], 'pk': self.kwargs['pk']}))    
        return super().dispatch(request, *args, **kwargs)
    template_name = 'monitoreo/estadoCuenta.html'
    model = detalleVenta

    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        idp = self.kwargs.get('idp', None)
        id = self.kwargs.get('pk', None)
        estado = estadoCuenta.objects.get(detalleVenta=id)
        cuotas = cuotaEstadoCuenta.objects.filter(estadoCuenta=estado)
        pagosm = pago.objects.filter(pagoMantenimiento__numeroCuotaEstadoCuenta__estadoCuenta=estado)
        pagosp = pago.objects.filter(prima__detalleVenta=id)
        sumPrima=0
        sumMantenimiento=0
        sumRecargoMantenimiento=0
        for pagoObject in pagosp:
            if pagoObject.prima !=None:
                sumPrima+=pagoObject.monto
        for pagoObject in pagosm:
            sumMantenimiento+=pagoObject.pagoMantenimiento.mantenimiento
        for pagoObject in pagosm:
            sumRecargoMantenimiento+=pagoObject.pagoMantenimiento.recargoMtto

        context['idp'] = idp
        context['id'] = id  
        context['pagosm'] = pagosm
        context['sumPrima'] = sumPrima
        context['sumMantenimiento'] = sumMantenimiento
        context['sumRecargoMantenimiento'] = sumRecargoMantenimiento
        return context

class EstadoCuentaReporte(TemplateView):

    def get_context_data(self, **kwargs):
        context=super().get_context_data(**kwargs)
        idp = self.kwargs.get('idp', None) 
        id = self.kwargs.get('pk', None)
        context['idp'] = idp
        context['id'] = id
        return context   

    def get(self,request,*args,**kwargs):
        datos = pagoMantenimiento.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Estado de Cuenta"
        ws.row_dimensions[6].height = 50.25
        ws.column_dimensions['C'].width = 12
        fuente = Font(size=9,bold=True)
        bordes = borders.Side(style = None, color = 'FF000000', border_style = 'thin')
        borde = Border(left = bordes, right = bordes, bottom = bordes, top = bordes)
        alineacion = Alignment(horizontal="center",vertical="center")
        col = 3
        celdas=26      

        ws.merge_cells('C2:AB3')
        ws.merge_cells('N5:R5')
        ws.merge_cells('S5:Y5')
        ws.merge_cells('AA5:AB5')
        ws['AA5'] = "Capital"
        ws['N5'] = "Intereses"
        ws['S5'] = "Otros"
        ws['N5'].alignment = alineacion
        ws['S5'].alignment = alineacion
        ws['N5'].border = borde
        ws['S5'].border = borde
        ws['AA5'].border = borde
        ws['AA5'].alignment = alineacion        
        ws['C2'] = "Estado de Cuenta"
        ws['C2'].font = Font(size=20,bold=True)
        ws['C2'].alignment = alineacion
        ws['C6'] = "Fecha de pago\ns/escritura"
        ws['D6'] = "No. Cuota"
        ws['E6'] = "Fecha Pago"
        ws['F6'] = "No. Rec."
        ws['G6'] = "Referencia"
        ws['H6'] = "Dìas interes"
        ws['I6'] = "Dìas mora\ncapital"
        ws['J6'] = "DESCUENTO\nPRONTO PAGO"
        ws['K6'] = "Pago total"
        ws['L6'] = "Tasa %"
        ws['M6'] = "Mora %"
        ws['N6'] = "Interés\nCorriente"
        ws['O6'] = "Intres Mora"
        ws['P6'] = "Subtotal\nIntereses"
        ws['Q6'] = "Pagados"
        ws['R6'] = "Pendientes"
        ws['S6'] = "Comisión"
        ws['T6'] = "Mtto"
        ws['U6'] = "Recargo"
        ws['V6'] = "Otros"
        ws['W6'] = "Subtotal"
        ws['X6'] = "Pagados"
        ws['Y6'] = "Pendientes"
        ws['Z6'] = "Total pagados"
        ws['AA6'] = "Abono"
        ws['AB6'] = "Saldo"
        col = 3
        for f in range(celdas):
            ws.cell(row=7,column=col).border = borde
            col +=1
        cont = 8
        nPagos = 1
        for q in datos:
            cuotaEstado = cuotaEstadoCuenta.objects.get(id = q.numeroCuotaEstadoCuenta_id)
            pagos = pago.objects.get(pagoMantenimiento_id = q.numeroReciboMantenimiento)
            ws.cell(row=cont,column=3).value 
            ws.cell(row=cont,column=4).value
            ws.cell(row=cont,column=5).value = pagos.fechaPago
            ws.cell(row=cont,column=6).value = q.numeroReciboMantenimiento
            if pagos.tipoPago == 2:
                ws.cell(row=cont,column=7).value = pagos.referencia
            else:
                ws.cell(row=cont,column=7).value = "Pago en Efectivo"
            ws.cell(row=cont,column=8).value 
            ws.cell(row=cont,column=9).value 
            ws.cell(row=cont,column=10).value 
            ws.cell(row=cont,column=11).value 
            ws.cell(row=cont,column=12).value 
            ws.cell(row=cont,column=13).value 
            ws.cell(row=cont,column=14).value 
            ws.cell(row=cont,column=15).value 
            ws.cell(row=cont,column=16).value 
            ws.cell(row=cont,column=17).value 
            ws.cell(row=cont,column=18).value 
            ws.cell(row=cont,column=19).value 
            ws.cell(row=cont,column=20).value = pagos.monto
            ws.cell(row=cont,column=21).value 
            ws.cell(row=cont,column=22).value = q.montoOtros
            ws.cell(row=cont,column=23).value = pagos.monto + q.montoOtros
            ws.cell(row=cont,column=24).value 
            ws.cell(row=cont,column=25).value 
            ws.cell(row=cont,column=26).value 
            ws.cell(row=cont,column=27).value
            ws.cell(row=cont,column=28).value
            col = 3
            for f in range(celdas):
                ws.cell(row=cont,column=col).border = borde
                #ws.cell(row=cont,column=col).number_format = '0.00'
                col+=1
            cont += 1
            nPagos += 1

        ws.cell(row=cont,column=3).value
        ws.cell(row=cont,column=5).value 
        ws.cell(row=cont,column=7).value = "Total"
        ws.cell(row=cont,column=10).value
        ws.cell(row=cont,column=11).value 
        ws.cell(row=cont,column=14).value 
        ws.cell(row=cont,column=15).value 
        ws.cell(row=cont,column=16).value 
        ws.cell(row=cont,column=17).value 
        ws.cell(row=cont,column=18).value 
        ws.cell(row=cont,column=19).value 
        ws.cell(row=cont,column=20).value 
        ws.cell(row=cont,column=21).value 
        ws.cell(row=cont,column=22).value 
        ws.cell(row=cont,column=23).value 
        ws.cell(row=cont,column=24).value 
        ws.cell(row=cont,column=25).value 
        ws.cell(row=cont,column=26).value 
        ws.cell(row=cont,column=27).value
        ws.cell(row=cont,column=28).value = ws.cell(row=cont-1,column=28).value
        col = 3
        for f in range(celdas):
            ws.cell(row=cont,column=col).border = borde
            ws.cell(row=cont,column=col).font = fuente
            ws.cell(row=6,column=col).font = fuente
            ws.cell(row=6,column=col).alignment = alineacion
            ws.cell(row=6,column=col).border = borde
            col+=1

        nombre_archivo = "EstadoCuenta.xlsx"
        response = HttpResponse()
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"]= contenido
        wb.save(response)
        return response